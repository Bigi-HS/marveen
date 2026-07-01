#!/usr/bin/env python3
"""
dubscript-to-xlsx.py -- Convert 16-column dubscript CSV to formatted Excel workbook.

Features:
- Header row frozen + bold
- Per-character background color (KARAKTER column)
- Conditional red fill: ON + syllable diff >20%
- TAKE column: yellow fill for RETAKE
- KAM column: color-coded badge (ON=red, OC=yellow, OFF=grey)
- Column widths auto-fitted
- Filter view on header row

Usage:
    python3 dubscript-to-xlsx.py \
        --csv store/spaceking-ep1/spaceking-ep1-template-v1.csv \
        --out store/spaceking-ep1/spaceking-ep1-template-v1.xlsx \
        [--sheet "Space King ep1"]
"""

import argparse
import csv
import sys

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import Rule, FormulaRule
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Character colors (pastel hex, fill)
CHAR_COLORS = {
    "Kapitány":   "BDD7EE",  # light blue
    "Dühöngő":    "FCE4D6",  # light orange
    "Bryce":      "E2EFDA",  # light green
    "Mogyi":      "EAD1DC",  # light purple
    "Marine":     "FFF2CC",  # light yellow
    "Koponya":    "D9D9D9",  # light grey
    "Pók":        "F4CCCC",  # light red
    "Robothang":  "CFE2F3",  # pale blue
    "Űrkirály":   "D9EAD3",  # pale green
    "_default":   "FFFFFF",
}

KAM_COLORS = {
    "ON":  "FF0000",  # red
    "OC":  "FFD700",  # gold/yellow
    "OFF": "B0B0B0",  # grey
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

COL_WIDTHS = {
    "GLOBAL_ID":      11,
    "#":               5,
    "TC_IN":          12,
    "TC_OUT":         12,
    "DUR":             7,
    "KAMERA":          8,
    "KARAKTER":       14,
    "SZOT_HU":         8,
    "SZOT_EN":         8,
    "MAGYAR_SZOVEG":  45,
    "DIREKCIO":       20,
    "INT":             6,
    "KIEJTES":        14,
    "ANGOL_SZOVEG":   40,
    "FORD_MEGJEGYZES": 30,
    "TAKE":            9,
}


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--csv", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--sheet", default="Dubscript")
    args = ap.parse_args()

    with open(args.csv, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    if not rows:
        print("ERROR: empty CSV", file=sys.stderr)
        return 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet

    # Header row
    for col_idx, field in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        width = COL_WIDTHS.get(field, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    col_map = {f: i + 1 for i, f in enumerate(fieldnames)}
    kam_col = col_map.get("KAMERA")
    char_col = col_map.get("KARAKTER")
    hu_col = col_map.get("SZOT_HU")
    en_col = col_map.get("SZOT_EN")
    take_col = col_map.get("TAKE")
    magyar_col = col_map.get("MAGYAR_SZOVEG")

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        char = row.get("KARAKTER", "").strip()
        kam = row.get("KAMERA", "").strip().upper()
        take = row.get("TAKE", "").strip().upper()

        char_fill = make_fill(CHAR_COLORS.get(char, CHAR_COLORS["_default"]))

        for col_idx, field in enumerate(fieldnames, start=1):
            val = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=(field == "MAGYAR_SZOVEG"), vertical="top")

            # Character color on KARAKTER column only
            if col_idx == char_col:
                cell.fill = char_fill

            # KAM badge color
            if col_idx == kam_col and kam in KAM_COLORS:
                cell.fill = make_fill(KAM_COLORS[kam])
                cell.font = Font(bold=True, color="FFFFFF" if kam == "ON" else "000000")
                cell.alignment = Alignment(horizontal="center")

            # TAKE: yellow for RETAKE
            if col_idx == take_col and take == "RETAKE":
                cell.fill = make_fill("FFF2CC")

            # Magyar sor: larger font
            if col_idx == magyar_col:
                cell.font = Font(size=12)

        # Lip-sync warning: ON + syllable diff >20% -> red SZOT_HU cell
        if kam == "ON" and hu_col and en_col:
            try:
                hu_s = int(row.get("SZOT_HU", 0) or 0)
                en_s = int(row.get("SZOT_EN", 0) or 0)
                if en_s > 0 and abs(hu_s - en_s) / en_s > 0.20:
                    ws.cell(row=row_idx, column=hu_col).fill = make_fill("FF9999")
                    ws.cell(row=row_idx, column=en_col).fill = make_fill("FF9999")
            except (ValueError, ZeroDivisionError):
                pass

        # Row height for wrapped Magyar szoveg
        ws.row_dimensions[row_idx].height = 30

    ws.row_dimensions[1].height = 22

    wb.save(args.out)
    print(f"Saved: {args.out} ({len(rows)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
